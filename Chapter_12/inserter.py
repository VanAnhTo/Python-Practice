import openpyxl
import openpyxl.cell
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter, column_index_from_string


rowStartN = 3
rowInsertM = 2
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

maxRow = sheet.max_row
maxColumn = sheet.max_column

#print(sheet.max_row)
#print(sheet.max_column)
wb1 = openpyxl.load_workbook('exampleCopy.xlsx')
sheet1 = wb1.active

for row in range(maxRow):
    a = sheet['A'+ str(row)].value
    b = sheet['B'+ str(row)].value
    c = sheet['C'+ str(row)].value


