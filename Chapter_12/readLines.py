import os,shutil
import openpyxl
import sys
import openpyxl.cell
from numpy.distutils.system_info import numarray_info
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter, column_index_from_string

folderPath = 'E:\\Project\\Python_Project\\Python-Practice\\Chapter_12\\test\\'

file = 'E:\\Project\\Python_Project\\Python-Practice\\Chapter_12\\test\\test2.txt'
#fileContent = open(file)
'''
with open(file) as f:
    content = f.readlines()

print(content)
'''

excelFile = '1.xlsx'
wb = openpyxl.load_workbook(excelFile)
sheet = wb.get_sheet_by_name('Sheet')
'''
for fileName in os.listdir(folderPath):
    print(fileName)
    print(folderPath+fileName)
    with open(folderPath+fileName) as f:
        content = f.readlines()
        print(content)
'''

with open(folderPath) as f:
    content = f.readlines()
    #print(content)

print(len(content))


for j in range(1,len(content)):
    sheet.cell(row=j,column=1).value = content[j]
    if j == len(content):
        sheet.cell(row=j, column=1).value = content[j+1]        
