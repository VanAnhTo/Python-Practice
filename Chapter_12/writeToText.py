import os,shutil
import openpyxl
import sys
import openpyxl.cell
from numpy.distutils.system_info import numarray_info
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter, column_index_from_string


excelFile = 'readLines.xlsx'
wb = openpyxl.load_workbook(excelFile)
sheet = wb.get_sheet_by_name('Sheet')

for j in range (1, sheet.max_column+1):

    textFile = open('Text_%s.txt'%j, 'w')

    for i in range (1, sheet.max_row+1):
        val = sheet.cell(row=i, column=j).value
        if val == None:
            continue
        print(j,i,val)
        textFile.write(val)

    print('-----------')
