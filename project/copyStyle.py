import openpyxl
#from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color, colors
from copy import copy

# File to be copied
wb = openpyxl.load_workbook("DataToComment.xlsx")  # Add file name
sheet = wb["Sheet1"] # Add Sheet name
sheet_paste = wb["Sheet2"]



# a  = sheet['B1']
# b = sheet['G3']
#
# #print(b.font, b.border)
# a._style = copy(b._style)

for r in range(1,5):
    for c in range (1,11):
        v = sheet.cell(row =r, column=c)
        p = sheet_paste.cell(row=r, column=c)
        p._style = v._style
        p.value = v.value

sheet_paste.merge_cells('B1:G1')
sheet_paste.merge_cells('H1:I1')
sheet_paste.merge_cells('J1:J2')
sheet_paste.merge_cells('J3:J4')

wb.save("DataToComment.xlsx")



# ft = Font(color=colors.RED, name='Arial', size=18)
# a.font = ft
# fl = PatternFill(fill_type= 'solid' , start_color='FFFFCC')
# a.fill = fl
# bor = Border(
#     left=Side(border_style='thin',color='000055'),
# right=Side(border_style='thin',color='000055'),
# top=Side(border_style='thin',color='000055'),
# bottom=Side(border_style='thin',color='000055'),
# diagonal=Side(border_style='thin',color='000055'),
# diagonal_direction=0,
# outline=Side(border_style='thin',color='000055'),
# vertical=Side(border_style='thin',color='000055'),
# horizontal=Side(border_style='thin',color='000055'))
# a.border = bor



'''
for row in sheet.rows:
    for cell in row:
        new_cell = new_sheet.cell(row=3,
                                  column=2, value=cell.value)
        if cell.has_style:
            cell.Font = new_cell.Font
            # new_cell.Border = cell.Border
            # new_cell.Fill = cell.Fill
            # new_cell.Number_format = cell.Number_format
            # new_cell.Protection = cell.Protection
            # new_cell.Alignment = cell.Alignment
'''


