import openpyxl
import openpyxl.cell
from openpyxl.styles import Font
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from copy import copy, deepcopy
import pyperclip


from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import datetime, time
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


#Getting Sheets from the Workbook
from openpyxl.utils import get_column_letter, column_index_from_string

wb = openpyxl.load_workbook('CommentToCodeBeamer.xlsx')
sheet_score = wb.get_sheet_by_name('Score')
sheet_comment = wb.get_sheet_by_name('Comment')


# print(sheet_comment['A3'].value)
# print(sheet_comment['A3'])

# for rowOfCellObjects in sheet_comment['A1':'J5']:
#     for cellObj in rowOfCellObjects:
#
#         print(cellObj.coordinate, cellObj.value)
#     print('--- END OF ROW ---')


'''
for row in range (2, sheet_score.max_row+1):
    srs_id = sheet_score['A' + str(row)].value
    #print(srs_id)
    for rowOfCellObjects in sheet_score['I'+str(row):'Q'+str(row)]:
        for cellObj in rowOfCellObjects:
            print(cellObj.coordinate, cellObj.value)
        print('--- END OF ROW ---')
'''

'''
a = copy(sheet_comment['A1':'J5'])

driver = webdriver.Firefox(executable_path=r'D:\AnhTo\Setup\Wed_driver\geckodriver.exe')
driver.maximize_window()
driver.get('http://abc/6780145')

username = driver.find_element_by_id("user")
username.send_keys("username")

pas = driver.find_element_by_id("password")
pas.send_keys("123567")

btnLogin = driver.find_element_by_css_selector(".login_button")
btnLogin.click()

time.sleep(2)

driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")

commentTab = driver.find_element_by_id("task-details-attachments-tab")
commentTab.click()

commentLinkText = driver.find_element_by_css_selector(".addCommentLink a")
commentLinkText.click()

iframe = driver.find_element_by_id("editor_new_ifr")
driver.switch_to.frame(iframe)


commentFeild = driver.find_element_by_id("tinymce")
commentFeild.send_keys(a)
'''


'''
c = sheet_comment['K13']
b = sheet_comment['B3']
c.value = b.value

c.fill = b.fill
c.number_format = b.number_format
c.protection = b.protection
c.alignment = b.alignment

wb.save('CommentToCodeBeamer.xlsx')
'''
