import openpyxl
import openpyxl.cell
from openpyxl.styles import Font

import os

os.system("CoytoClipboard.exe")


'''
#Getting Sheets from the Workbook
from openpyxl.utils import get_column_letter, column_index_from_string

wb = openpyxl.load_workbook('CommentToCodeBeamer.xlsx')
sheet_score = wb.get_sheet_by_name('Score')
sheet_comment = wb.get_sheet_by_name('Comment')

for row in range (2, sheet_score.max_row+1):
    srs_id = sheet_score['A' + str(row)].value
    #print(srs_id)
    for rowOfCellObjects in sheet_score['I'+str(row):'Q'+str(row)]:
        for cellObj in rowOfCellObjects:
            print(cellObj.coordinate, cellObj.value)


        print('--- END OF ROW ---')
'''
btnLogin = driver.find_element_by_css_selector(".login_button")
btnLogin.click()

time.sleep(2)

driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")

commentTab = driver.find_element_by_id("task-details-attachments-tab")
commentTab.click()

commentLinkText = driver.find_element_by_css_selector(".addCommentLink a")
commentLinkText.click()

frame = driver.find_element_by_id("editor_new_ifr")
driver.switch_to.frame(frame)
time.sleep(2)

commentField = driver.find_element_by_id("tinymce")
commentField.send_keys()
