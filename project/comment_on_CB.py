'''
This is source to do:
1. Open link of SRS base on the ID colume of the SRS scoring file.
2. Copy score from SRS scoring file and paste to template comment file
3. From template comment file, copy and paste to the Comment field of the SRS (link is opened in step 1).
4. Save the comment.
5. Open and do the same with the remaining ID on the SRS scoring file.
'''

import os
import time

import openpyxl
import logging
from selenium import webdriver
from selenium.webdriver.common.keys import Keys

from ICAS3_Comment_CB import my_config

logging.basicConfig(filename='logFile.txt',
                        level=logging.DEBUG,
                        format='%(asctime)s - %(levelname)s - %(message)s')

logging.debug('Start of program')

score_workbook = openpyxl.load_workbook(my_config.scoring_file)
sheet_score = score_workbook[my_config.sheet_score_name]

wb = openpyxl.load_workbook(my_config.comment_file)
sheet_comment = wb[my_config.sheet_comment_name]

numberOfId = sheet_score.max_row

driver = webdriver.Firefox(executable_path= my_config.firefox_geckodriver)
driver.maximize_window()
driver.get(my_config.code_beamer_link)


username = driver.find_element_by_id("user")
username.send_keys(my_config.username)

pas = driver.find_element_by_id("password")
pas.send_keys(my_config.password)

btnLogin = driver.find_element_by_css_selector(".login_button")
btnLogin.click()

time.sleep(2)

for r in range (2, numberOfId+1):
    srs_ID = sheet_score['A'+ str(r)].value
    logging.debug('Running SRS_ID: ' + str(srs_ID))
    driver.get(my_config.srs_link + str(srs_ID))
    time.sleep(2)

    for c in range(1, 10):
        v = sheet_score.cell(row=r, column=c + 8)  # Data copied with row increased every cover loop
        p = sheet_comment.cell(row=3, column=c)  # Data filled to the fixed row of the comment file row =3.
        p.value = v.value

    wb.save(my_config.comment_file)

    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")

    commentTab = driver.find_element_by_id("task-details-attachments-tab")
    commentTab.click()

    commentLinkText = driver.find_element_by_css_selector(".addCommentLink a")
    commentLinkText.click()

    frame = driver.find_element_by_id("editor_new_ifr")
    driver.switch_to.frame(frame)

    os.system(my_config.autoIT_copy_data_exe)

    commentField = driver.find_element_by_id("tinymce")
    commentField.send_keys(Keys.CONTROL, "v")

    #commentField.send_keys(Keys.CONTROL, "s")

    os.system(my_config.autoIT_close_excel)

    time.sleep(3)

driver.quit()

logging.debug('End of program')
