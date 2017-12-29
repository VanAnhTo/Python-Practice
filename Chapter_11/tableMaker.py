import openpyxl
import sys
import openpyxl.cell
from numpy.distutils.system_info import numarray_info
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter, column_index_from_string

num = ''.join(sys.argv[1:])

numberTable = int(num)

wb = openpyxl.Workbook()
sheet = wb.active
boldFont = Font(bold=True)

for i in range(2,numberTable+2):
    for j in range(2,numberTable+2):

        sheet.cell(row=i, column=1).value = i - 1
        sheet.cell(row=i, column=1).font = boldFont

        sheet.cell(row=1, column=j).value = j - 1
        sheet.cell(row=1, column=j).font = boldFont

        sheet.cell(row=i, column=j).value = (i-1)*(j-1)

wb.save('1.xlsx')





