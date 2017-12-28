#Installing the openpyxl Module
import openpyxl

import openpyxl.cell

#Getting Sheets from the Workbook
from openpyxl.utils import get_column_letter, column_index_from_string

'''
wb = openpyxl.load_workbook('example.xlsx')
wb.get_sheet_names()

sheet = wb.get_sheet_by_name('Sheet1')


wb = openpyxl.load_workbook('example.xlsx')
print(wb.get_sheet_names())

sheet = wb.get_sheet_by_name('Sheet3')
print(sheet)

print(type(sheet))
print(sheet.title)

anotherSheet = wb.active
print(anotherSheet)
'''
#Getting Cells from the Sheets

'''
print(sheet['A1'])
print(sheet['A1'].value)

c = sheet['B1']
print(c.value)

print('Row '+str(c.row) +', Column '+ c.column + ' is '+ c.value)
print(('Cell '+ c.coordinate + ' is '+ c.value))

d = sheet['C1'].value
print(d)

print(sheet.cell(row=1, column=2))
print(sheet.cell(row=1, column=2).value)


for i in range(1,8,2):
    print(i, sheet.cell(row=i, column=2).value)

print(sheet.max_row)
print(sheet.max_column)
'''

#Converting Between Column Letters and Numbers
'''
print(get_column_letter(1))
print(get_column_letter(5))
print(get_column_letter(sheet.max_column))

print(column_index_from_string('A'))
print(column_index_from_string('AA'))

#Getting Rows and Columns from the Sheets
tuple(sheet['A1':'C3'])

for rowOfCellObjects in sheet['A1':'C3']:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
    print('--- END OF ROW ---')

sheet = wb.active
print(list(sheet.columns)[1])

for cellObj in list(sheet.columns)[1]:
        print(cellObj.value)
'''

#Workbooks, Sheets, Cells

#Creating and Saving Excel Documents
'''
wb = openpyxl.Workbook()
wb.get_sheet_names()
sheet = wb.active
print(sheet.title)
sheet.title = 'Spam Bacon Eggs Sheet'
print(wb.get_sheet_names())
'''
'''
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.active
sheet.title = 'Spam Spam Spam'
wb.save('example_copy.xlsx')
'''

#Creating and Removing Sheets
'''
wb = openpyxl.Workbook()
print(wb.get_sheet_names())

wb.create_sheet()
print(wb.get_sheet_names())

wb.create_sheet(index=0, title='First Sheet')
print(wb.get_sheet_names())

wb.create_sheet(index=2, title='Middle Sheet')
print(wb.get_sheet_names())

wb.remove_sheet(wb.get_sheet_by_name('Middle Sheet'))
wb.remove_sheet(wb.get_sheet_by_name('Sheet1'))
print(wb.get_sheet_names())

#**Remember to call the save() method to save the changes after adding sheets to or removing sheets from the workbook.
'''

#Writing Values to Cells
'''
wb = openpyxl.Workbook()
sheet = wb.get_sheet_by_name('Sheet')
sheet['A1'] = 'Hello world!'
print(sheet['A1'].value)
'''

#Setting the Font Style of Cells

from openpyxl.styles import Font
'''
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
italic24Font = Font(size=24, italic=True)
sheet['A1'].font = italic24Font
wb.save('styled.xlsx')
'''
'''
wb = openpyxl.Workbook()
sheet = wb.get_sheet_by_name('Sheet')
fontObj1 = Font(name='Times New Roman', bold=True)
sheet['A1'].font = fontObj1
sheet['A1'] = 'Bold Times New Roman'

fontObj2 = Font(size=24, italic=True)
sheet['B3'].font = fontObj2
sheet['B3'] = '24 pt Italic'

wb.save('styles.xlsx')
'''

#Formulas
'''
wb = openpyxl.Workbook()
sheet = wb.active
sheet['A1'] = 200
sheet['A2'] = 300
sheet['A3'] = '=SUM(A1:A2)'
wb.save('writeFormula.xlsx')
'''

#Adjusting Rows and Columns
#Setting Row Height and Column Width
'''
wb = openpyxl.Workbook()
sheet = wb.active

sheet['A1'] = 'Tall row'
sheet['B2'] = 'Wide column'

sheet.row_dimensions[1].height = 70
sheet.column_dimensions['B'].width = 20
wb.save('dimensions.xlsx')
'''

#Merging and Unmerging Cells
'''
wb = openpyxl.Workbook()
sheet = wb.active
sheet.merge_cells('A1:D3')
sheet['A1'] = 'Twelve cells merged together.'
sheet.merge_cells('C5:D5')
sheet['C5'] = 'Two merged cells.'
wb.save('merged.xlsx')

#unmerge
wb = openpyxl.load_workbook('merged.xlsx')
sheet = wb.active
sheet.unmerge_cells('A1:D3')
sheet.unmerge_cells('C5:D5')
wb.save('merged.xlsx')
'''

#Freeze Panes
'''
wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.active
sheet.freeze_panes = 'A2'
wb.save('freezeExample.xlsx')
'''
    
#Charts
wb = openpyxl.Workbook()
sheet = wb.active
for i in range(1, 11):
    sheet['A' + str(i)] = i

refObj = openpyxl.chart.Reference(sheet, min_col=1, min_row=1, max_col=1, max_row=10)
seriesObj = openpyxl.chart.Series(refObj, title='First series')
chartObj = openpyxl.chart.BarChart()
chartObj.title = 'My Chart'
chartObj.append(seriesObj)
sheet.add_chart(chartObj, 'C5')
wb.save('sampleChart.xlsx')


