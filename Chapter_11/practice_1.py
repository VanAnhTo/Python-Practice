import openpyxl
import openpyxl.cell
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter, column_index_from_string

numberTable = 6

wb = openpyxl.Workbook()
sheet = wb.active
for i in range(1,7):
    for j in range(1,7):
        sheet.cell(row=i, column=j).value = i*j
wb.save('1.xlsx')





